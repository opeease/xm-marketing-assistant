import json
import os
import re
import time
from pathlib import Path
from typing import Any, Dict, List, Optional

from loguru import logger


def get_local_config(key: str, default: Any = None) -> Any:
    """获取本地配置(从 store.json)"""
    store_path = os.path.join(os.path.expanduser("~"), "dt-ai-helper", "store.json")
    try:
        if os.path.exists(store_path):
            data = json.loads(Path(store_path).read_text(encoding="utf-8"))
            return data.get(key, default)
    except Exception:
        pass
    return default


def set_local_config(key: str, value: Any):
    """设置本地配置"""
    store_path = os.path.join(os.path.expanduser("~"), "dt-ai-helper", "store.json")
    try:
        os.makedirs(os.path.dirname(store_path), exist_ok=True)
        data = {}
        if os.path.exists(store_path):
            data = json.loads(Path(store_path).read_text(encoding="utf-8"))
        data[key] = value
        Path(store_path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        logger.error(f"set_local_config 失败: {e}")


def get_wait_time(config_key: str = None) -> int:
    """获取等待时间(随机化防封)"""
    import random
    base = 2
    if config_key:
        try:
            conf = json.loads(get_local_config(config_key, "{}"))
            base = conf.get("waitTime", 2)
        except Exception:
            pass
    return base + random.randint(0, 2)


def is_pyinstaller() -> bool:
    """判断是否 PyInstaller 打包"""
    return False


def open_program(path: str):
    """打开程序"""
    if path and os.path.exists(path):
        os.startfile(path)


def close_open():
    """关闭操作"""
    pass


def init_open_status():
    """初始化开启状态"""
    pass


def copy_to_clipboard_windows(text: str):
    """复制到剪贴板"""
    import pyperclip
    pyperclip.copy(text)


def copy_files_to_clipboard(file_paths: List[str]):
    """复制文件到剪贴板"""
    import subprocess
    if not file_paths:
        return
    paths = ";".join(f'"{p}"' for p in file_paths if os.path.exists(p))
    if paths:
        subprocess.run(
            ["powershell", "-Command", f"Set-Clipboard -Path {paths}"],
            capture_output=True,
            timeout=10,
        )


def send_windows_systemp_notification(title: str, message: str):
    """发送 Windows 系统通知"""
    try:
        from plyer import notification
        notification.notify(title=title, message=message, timeout=5)
    except Exception:
        pass


def batch_import_contents(file_path: str, header_name: str = "账号",
                          is_tiktok_link: bool = False) -> list:
    """从文件批量导入内容"""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".txt":
        lines = Path(file_path).read_text(encoding="utf-8").splitlines()
        return [{"content": l.strip(), "source": file_path} for l in lines if l.strip()]
    elif ext == ".xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(values_only=True))
            col_idx = None
            for i, h in enumerate(header_row):
                if h and str(h).strip() == header_name:
                    col_idx = i
                    break
            if col_idx is None:
                raise KeyError(f"未找到'{header_name}'列")
            results = []
            for row in ws.iter_rows(values_only=True):
                val = row[col_idx]
                if val and str(val).strip():
                    results.append({"content": str(val).strip(), "source": file_path})
            return results
        except ImportError:
            logger.error("需要安装 openpyxl: pip install openpyxl")
            return []
    return []


def with_log_context(func):
    """带日志上下文的装饰器"""
    return func
